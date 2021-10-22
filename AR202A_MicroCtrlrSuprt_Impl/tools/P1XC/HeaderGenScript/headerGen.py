#! python2
# NAME: headerGen.py script
# VERSION: v1.3 - 3/14/2017
# DESCRIPTION: Please see Word document "HeaderGen.docx" provided with this script 

import sys
import re
import time
from openpyxl import load_workbook
import warnings


# -----------------------------------------------------------------------------------------------
# CLASSES
# -----------------------------------------------------------------------------------------------

#
# NAME: ExcelRow
# DESCRIPTION: Class ExcelRow is the container for a single row of the input Excel file. Each argument to the __init__
# function corresponds to a column in the Excel, identified by letter.
#
class ExcelRow(object):
    def __init__(self, _a, _b, _c, _d, _e, _f, _g, _h, _i, _j, _k, _l, _m, _n, _o, _p, _q, _r, _s):
        self.group = _a
        self.address = _b
        self.name = _c
        self.size_1 = _d
        self.size_8 = _e
        self.size_16 = _f
        self.size_32 = _g
        self.bp7 = _h  # bit position 7
        self.bp6 = _i  # bit position 6
        self.bp5 = _j  # bit position 5
        self.bp4 = _k  # bit position 4
        self.bp3 = _l  # bit position 3
        self.bp2 = _m  # bit position 2
        self.bp1 = _n  # bit position 1
        self.bp0 = _o  # bit position 0
        self.access = _p
        self.start = _q
        self.size = _r
        self.read_only = _s

    def __repr__(self):
        return "\n<name: %s, address: %s>" % (self.name, self.address)


#
# NAME: OutputStrings
# DESCRIPTION: Class OutputStrings contains strings to be printed to the output file for each
# of the major components of the file: the prefix, the groups, and the suffix. The groups are
# described in class OutputStringGroup, below
#
class OutputStrings(object):
    def __init__(self, prefix, groups, suffix):
        self.prefix = prefix
        self.groups = groups
        self.suffix = suffix


#
# NAME: OutputStringGroup
# DESCRIPTION: Class OutputStringGroups contains strings to eb printed to the output file for each
# of the register groups as defined in the config file. Each register group consists of a structure for the
# base types ("base"), one for the bit accesses ("bits"), one for the access macros ("macros") and one for the
# compiler binding pragmas ("bind").
#
class OutputStringGroup(object):
    def __init__(self, base, bits, macros, bind, macros_addr):
        self.base = base
        self.bits = bits
        self.macros = macros
        self.bind = bind
        self.macros_addr = macros_addr


#
# NAME: RegSizeHisto
# DESCRIPTION: Class RegSizeHisto stores a histogram of the register access sizes as parsed from the Excel file.
# This is used to determine not only what to print to the file in terms of access struct members, but how to
# contonue parsing the file as well.
#
class RegSizeHisto(object):
    def __init__(self, size32, size16, size8, size1, size_blank):
        self.blank = size_blank
        self.size1 = size1
        self.size8 = size8
        self.size16 = size16
        self.size32 = size32

    def __repr__(self):
        return "<blank: %s, 1: %s, 8: %s, 16: %s, 32: %s>" % (
            self.blank, self.size1, self.size8, self.size16, self.size32)


# -----------------------------------------------------------------------------------------------
# SORTING KEYS
# -----------------------------------------------------------------------------------------------
#
# NAME: sort_by_address
# DESCRIPTION: performs a sort based on the address field of the object ExcelRow
#
def sort_by_address(excel_row):
    return excel_row.address


# -----------------------------------------------------------------------------------------------
# FILTERS
# -----------------------------------------------------------------------------------------------
#
# NAME: filter_base_types
# DESCRIPTION: Filter filter_base_types returns true if the object ExcelRow corresponds to a "base type", meaning
# an 8, 16, or 32-bit raw access instead of a bitfield access.
# RETURN: 0: not a base type, 1: base type
#
def filter_base_types(x):
    # look for entries that have no base type selected - these are base type definitions
    if (x.size_8 is None) and (x.size_16 is None) and (x.size_32 is None):
        return 0
    else:
        return 1


#
# NAME: filter_bit_types
# DESCRIPTION: Filter filter_bit_types returns true if the object ExcelRow corresponds to a "bit type", meaning
# it is accessing a named portion of the register.
# RETURN: 0, not a bit type, 1: bit type
#
def filter_bit_types(x):
    # look for entries that have bit size identified - these are bit definitions
    if (x.size is not None) or ((x.size_1 is not None) and (x.size_8 is None)):
        return 1
    else:
        return 0


#
# NAME: filter_address_match
# DESCRIPTION: Filter filter_address_match returns true if member 'address' of object ExcelRow matches the
# one passed into the filter function
# RETURN: 0: address does not match, 1: address matches
#
def filter_address_match(x, address):
    return x.address == address


#
# NAME: filter_address_match_range
# DESCRIPTION: Filter filter_address_match_range returns true if the member 'address' of object ExcelRow is between the
# provided start address and the provided delta
# RETURN: 0: address is outside of range, 1: address is in range
#
def filter_address_match_range(x, address, delta):
    if (int(x.address, 16) >= int(address, 16)) and (int(x.address, 16) < (int(address, 16) + delta)):
        return 1
    else:
        return 0


# -----------------------------------------------------------------------------------------------
# FUNCTIONS
# -----------------------------------------------------------------------------------------------

#
# NAME: parse_config_file
# DESCRIPTION: Function parse_config_file reads the configuration file as specified and loads all pertinent information
# into a structure for use by the rest of the program
# INPUTS: config_file_name: name of the configuration file to open
# RETURN: dictionary config_data, populated with configuration file information
#
def parse_config_file(config_file_name):
    print "parsing config file...\n"
    config_data = {}
    prefix = ""
    suffix = ""
    groups = []
    skip_list = []
    config_file_valid_lines = []

    config_file_object = open(config_file_name, "r")

    # strip blank lines
    config_file_lines = filter(None, (line.rstrip() for line in config_file_object))

    # done with file object - data lives in config_file_lines now
    config_file_object.close()

    for line in config_file_lines:
        if not line.startswith("#"):
            config_file_valid_lines.append(line)

    config_data["use_module_name"] = False
    # if the user wants to use group names as the output file names, find out now
    for line in config_file_valid_lines:
        if "use_module_name = true" in line or "use_module_name = True" in line:
            config_data["use_module_name"] = True
            config_file_valid_lines.remove(line)
        if "use_module_name = false" in line or "use_module_name = False" in line:
            config_data["use_module_name"] = False
            config_file_valid_lines.remove(line)

    # if the user wants to generate register address macros, find out now
    for line in config_file_valid_lines:
        if "gen_address_macros = true" in line or "gen_address_macros = True" in line:
            config_data["gen_address_macros"] = True
            config_file_valid_lines.remove(line)
        if "gen_address_macros = false" in line or "gen_address_macros = False" in line:
            config_data["gen_address_macros"] = False
            config_file_valid_lines.remove(line)

    # parse configuration file - only store characters after the identification token ("[prefix]", etc...)
    for line in config_file_valid_lines:
        line = line.lstrip()

        if "[prefix]" in line:
            prefix += line[9:] + "\n"

        if "[suffix]" in line:
            suffix += line[9:] + "\n"

        if "[groups]" in line and (config_data["use_module_name"] is False):
            # normal group
            group_name, group_string = line[9:].split("=")
            temp_array = [group_name.strip()]
            for i in group_string.split(","):
                temp_array.append(i.strip())
            groups.append(temp_array)

        if "[skip]" in line:
            skip_list.append(line[7:])
        elif not line.startswith("["):
            # lines without an identification token can just be split on the "=" and stored into the dict
            key, value = line.split("=")
            config_data[key.strip()] = value.strip()

    if skip_list:
        config_data["skip"] = skip_list
        groups.append(["SKIP", ""])

    config_data["prefix"] = prefix
    config_data["groups"] = groups
    config_data["suffix"] = suffix

    # input file should have a micro identifier in the name - find it
    micro_id_reg_pattern = re.compile(r"dr7f70....")
    device_name = micro_id_reg_pattern.search(config_data["input_file_name"]).group(0)
    config_data["device_name"] = device_name

    # concatenate output file name
    config_data["output_file"] = config_data["output_file_location"] + device_name

    return config_data


# ----------------------------------------------------------
#
# NAME: load_excel
# DESCRIPTION: Function load_excel pulls the input Excel file into a list of ExcelRow objects for consumption
# by the program. This can take quite a while as the file can be very large.
# NOTE: This function uses the read_only attribute of the openpyxl package, which requires the program to read
# cells programmatically from top to bottom and then from left to right. Reading in another pattern will cause
# delays
# INPUTS: config_data: the configuration data dictionary
#         excel_rows: list of ExcelRow objects
# RETURN: maximum number of rows in the Excel File
#
#
def load_excel(config_data, excel_rows):
    print "loading excel file into memory...\n"
    wb = load_workbook(filename=config_data["input_file_name"], read_only=True)
    wbs = wb[config_data["tab_name"]]

    max_row = wbs.max_row
    i = 0
   
	
    for single_row in wbs.rows:
        temp_array = []

        for cell in single_row:
	    try:
                newval = cell.value.strip()
	    except:
		newval = cell.value
            temp_array.append(newval)

        if temp_array[0] == u'Module Name' or temp_array[0] is None:
            continue

		
        read_only = True
        read_only_array = temp_array[3:16]
        for x in read_only_array:
            if (x == u'E') or (x == u'W'):
                read_only = False
                break
		
        # Address field should be forced to unicode. During import, numbers with non-numeric hex characters
        # (a-f) are assumed to be unicode, but hex number without non-numeric characters (0x12345678) are
        # not unicode, which breaks the parsing.
        temp_row = ExcelRow(temp_array[0],  # group
                            unicode(temp_array[1]),  # address
                            "_" + temp_array[2],  # name
                            temp_array[3],  # 1 bit signal
                            temp_array[4],  # 8 bit signal
                            temp_array[5],  # 16 bit signal
                            temp_array[6],  # 32 bit signal
                            temp_array[7],  # bit position #7
                            temp_array[8],  # bit position #6
                            temp_array[9],  # bit position #5
                            temp_array[10],  # bit position #4
                            temp_array[11],  # bit position #3
                            temp_array[12],  # bit position #2
                            temp_array[13],  # bit position #1
                            temp_array[14],  # bit position #0
                            temp_array[15],  # access type
                            temp_array[16],  # signal start
                            temp_array[17],  # signal size
                            read_only)
        

        # None and unicode empty string ('') are used interchangeably in the Excel file, make them all None
        if temp_row.group is unicode(''):
            temp_row.group = None
        if temp_row.address is unicode(''):
            temp_row.address = None
        if temp_row.bp0 is unicode(''):
            temp_row.bp0 = None
        if temp_row.bp1 is unicode(''):
            temp_row.bp1 = None
        if temp_row.bp2 is unicode(''):
            temp_row.bp2 = None
        if temp_row.bp3 is unicode(''):
            temp_row.bp3 = None
        if temp_row.bp4 is unicode(''):
            temp_row.bp4 = None
        if temp_row.bp5 is unicode(''):
            temp_row.bp5 = None
        if temp_row.bp6 is unicode(''):
            temp_row.bp6 = None
        if temp_row.bp7 is unicode(''):
            temp_row.bp7 = None
        if temp_row.group is unicode(''):
            temp_row.group = None
        if temp_row.name is unicode(''):
            temp_row.name = None
        if temp_row.size is unicode(''):
            temp_row.size = None
        if temp_row.size_1 is unicode(''):
            temp_row.size_1 = None
        if temp_row.size_16 is unicode(''):
            temp_row.size_16 = None
        if temp_row.size_32 is unicode(''):
            temp_row.size_32 = None
        if temp_row.size_8 is unicode(''):
            temp_row.size_8 = None
        if temp_row.start is unicode(''):
            temp_row.start = None

        excel_rows.append(temp_row)

        # print status - use flush and '\r' to wipe out the previous row and print on top of it to create a status bar
        sys.stdout.flush()
        sys.stdout.write('\rparsing excel file for valid entries: {:.0%}'.format(float(i) / float(max_row)))
        i += 1

    return max_row


#
# NAME: find_module_names
# DESCRIPTION: iterate through the Excel file to find all unique module names, which will be used as groups
# for the output files.
# INPUTS: excel_rows: the unsorted list of Excel row objects
#
def find_module_names(excel_rows, config_data):
    unique_module_name = []
    temp_array = []
    for line in excel_rows:
        if line.group not in unique_module_name:
            unique_module_name.append(line.group)
            temp_array = [str(line.group), ""]
            config_data["groups"].append(temp_array)


#
# NAME: process_register
# DESCRIPTION: Function process_register is the primary function of this script. This function accepts the entire list
# of address sorted Excel objects, as well as an address to match. Once it finds a match, it will identify the size
# of the register, collect all Excel row objects that within the address range, and parse them for data. It will output
# the formatted header file strings into the appropriate object and then exit. This function is called for *register*
# but not necessarily for each row in the Excel list, as they are consumed in groups. Because this function is run
# so frequently, care must be taken to minimize content.
# INPUTS: address_for_match: the first address of the register to find and process
#         excel_rows_sorted: the sorted list of Excel row objects
#         output_strings: collection of strings to print to the output files; the primary output of this script
#         config_data: configuration data object
#
def process_register(address_for_match, excel_rows_sorted, output_strings, config_data):
    t0 = time.clock()

    address_delta = 0
    base_type_struct_name = ""
    output_group_name = "DEFAULT"
    base_type_string = "typedef union\n{\n"
    base_type_macro_string = ""

    # use list comprehension to collect all row objects that match the provided address.
    filtered_list = [x for x in excel_rows_sorted if filter_address_match(x, address_for_match)]

    # TODO: this should be unnecessary as we are adjusting address_for_match to a known good address at function end
    if filtered_list:
        # sort based on 32, 16, and then 8 bit entries - the largest entry will be on the bottom of the list
        size_filtered_list = sorted(filtered_list, key=lambda y: (y.size_32, y.size_16, y.size_8))

        # find out how large the largest row is, this will define the number of addresses that are a part of this
        # register, and also what the name of the register is
        biggest_type = size_filtered_list[-1]
        if biggest_type.size_32 is not None:
            address_delta = 4
            base_type_struct_name = biggest_type.name
        elif biggest_type.size_16 is not None:
            address_delta = 2
            base_type_struct_name = biggest_type.name
        elif biggest_type.size_8 is not None:
            address_delta = 1
            base_type_struct_name = biggest_type.name

        # TODO: optimize this searching for output group name - think about how 'skip' should work!
        # TODO: we need to handle the "use_module_name" portion of this code!
        # see if this register name matches any of the pre-defined groups from the config file, otherwise we will
        # use the default group
        if 'skip' in config_data:
            for skip_address in config_data["skip"]:
                if (int(address_for_match, 16) <= int(skip_address, 16)) and \
                        (int(address_for_match, 16) + address_delta > int(skip_address, 16)):
                    # store into skip category
                    output_group_name = "SKIP"

        # if we're NOT skipping, find the correct group
        if output_group_name is not "SKIP":
            if config_data["use_module_name"] is False:
                for temp_group_list in config_data["groups"]:
                    for temp_group_name in temp_group_list[1:]:
                        if temp_group_name is not "" and base_type_struct_name.startswith(temp_group_name):
                            output_group_name = temp_group_list[0]
            else:
                output_group_name = filtered_list[0].group

        # filter list to match a range of addresses
        # result: all entries matching a range of addresses with a max range of 4 (32-bit register)
        filtered_list = [x for x in excel_rows_sorted if filter_address_match_range(x,
                                                                                    address_for_match,
                                                                                    address_delta)]

        # see if we can apply the 'const' prefix here
        temp_const_prefix = ", __READ, "
        for row in filtered_list:
            if row.read_only is False:
                temp_const_prefix = ", __READ_WRITE, "

        # create binding string (only GHS is supported right now)
        binding_string = "__IOREG(" + base_type_struct_name[1:] + "_BASE, 0x" + address_for_match + \
                         temp_const_prefix + base_type_struct_name[1:] + "_t);\n"

        if config_data["gen_address_macros"] is True:
            # create address mapping string
            address_macro_string = "#define " + base_type_struct_name + "_ADDR (0x" + address_for_match + "UL)\n"
            output_strings.groups[output_group_name].macros_addr += address_macro_string

        # no further processing needed for compiler binding strings, save to output_strings variable
        output_strings.groups[output_group_name].bind += binding_string

        base_type_count = RegSizeHisto(0, 0, 0, 0, 0)

        # filter list to show all base types
        # result: all entries that are just byte/half-word/word accesses
        filtered_list_base_types_only = filter(filter_base_types, filtered_list)

        for row in filtered_list_base_types_only:
            if row.size_32 is not None:
                base_type_macro_string += "#define " + base_type_struct_name[1:] + " " + base_type_struct_name[1:] + "_BASE." + \
                                          config_data["base_union_name_long"] + "\n"
                base_type_string += "    " + config_data["base_type_long"] + " " + \
                                    config_data["base_union_name_long"] + ";\n"

            elif row.size_16 is not None:
                base_type_count.size16 += 1

            elif row.size_8 is not None:
                base_type_count.size8 += 1

        if base_type_count.size16 == 2:
            base_type_string += "    " + config_data["base_type_short"] + " " + \
                                config_data["base_union_name_short"] + "[2];\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + "L " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_short"] + "[L]\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + "H " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_short"] + "[H]\n"
        elif base_type_count.size16 == 1:
            base_type_string += "    " + config_data["base_type_short"] + " " + \
                                config_data["base_union_name_short"] + ";\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + " " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_short"] + "\n"

        if base_type_count.size8 == 4:
            base_type_string += "    " + config_data["base_type_byte"] + " " + \
                                config_data["base_union_name_byte"] + "[4];\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + "LL " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_byte"] + "[LL]\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + "LH " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_byte"] + "[LH]\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + "HL " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_byte"] + "[HL]\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + "HH " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_byte"] + "[HH]\n"
        elif base_type_count.size8 == 3:
            print "\n24 bit register near address 0x" + address_for_match + "??"
            exit()
        elif base_type_count.size8 == 2:
            base_type_string += "    " + config_data["base_type_byte"] + " " + \
                                config_data["base_union_name_byte"] + "[2];\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + "L " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_byte"] + "[L]\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + "H " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_byte"] + "[H]\n"
        elif base_type_count.size8 == 1:
            base_type_string += "    " + config_data["base_type_byte"] + " " + \
                                config_data["base_union_name_byte"] + ";\n"
            base_type_macro_string += "#define " + base_type_struct_name[1:] + " " + base_type_struct_name[1:] + "_BASE." + \
                                      config_data["base_union_name_byte"] + "\n"

        # filter list to show all individual register "bit types"
        # result: all entries that are just individual registers or "bits"
        filtered_list_bit_types_only = filter(filter_bit_types, filtered_list)

        if filtered_list_bit_types_only:
            # add the bit type to the base type union
            # see if we can apply the 'const' prefix here
            temp_const_prefix = "const "
            for row in filtered_list_bit_types_only:
                if row.read_only is False:
                    temp_const_prefix = ""

            base_type_string += "    " + temp_const_prefix + base_type_struct_name[1:] + "Bits_t " + \
                                config_data["base_union_name_bits"] + ";\n"

        base_type_string += "} " + base_type_struct_name[1:] + "_t;\n\n"

        # print base_type_string
        output_strings.groups[output_group_name].base += base_type_string
        output_strings.groups[output_group_name].macros += base_type_macro_string

        # extra computation needed if there are bit types defined in the register
        if filtered_list_bit_types_only:

            # start building the bit sub-structure
            bit_type_string = "typedef struct\n{\n"
            bit_type_macro_string = ""
            bit_list = []

            # determine what the base type of the bitfield will be
            if address_delta == 4:
                temp_bitfield_type = config_data["base_type_long"]
            elif address_delta == 2:
                temp_bitfield_type = config_data["base_type_short"]
            else:
                temp_bitfield_type = config_data["base_type_byte"]

            # for each byte in the register:
            for i in range(address_delta):
                # create a byte of single bit dummy fields to be overwritten later
                for j in range(8):
                    bit_list.append("    " + temp_bitfield_type + " padding" + str(j + (8 * i)) + " : 1;\n")

                # for single bit fields, replace dummy entries with named entries
                for bit in filtered_list_bit_types_only:
                    if (int(bit.address, 16) == (int(address_for_match, 16) + i)) and bit.size is None:
                        # replace the appropriate rows with this signal
                        bit_position_list = [bit.bp0,
                                             bit.bp1,
                                             bit.bp2,
                                             bit.bp3,
                                             bit.bp4,
                                             bit.bp5,
                                             bit.bp6,
                                             bit.bp7]

                        for k in range(8):
                            if bit_position_list[k] is not None:
                                if bit.read_only is True:
                                    bit_list[k + (8 * i)] = "    const " + temp_bitfield_type + " " + bit.name \
                                                            + " : 1;\n"
                                else:
                                    bit_list[k + (8 * i)] = "    " + temp_bitfield_type + " " + bit.name + " : 1;\n"

            # for multibit fields, replace dummy entries with named entries
            for bit in filtered_list_bit_types_only:
                if bit.size is not None:
                    for j in range(int(bit.size)):
                        temp1 = int(bit.address, 16)
                        temp2 = int(address_for_match, 16)
                        temp3 = int(bit.start)
                        if bit.read_only is True:
                            bit_list[((temp1 - temp2) * 8) + j + temp3] = "    const " + temp_bitfield_type + " " + \
                                                                          bit.name + " : 1;\n"
                        else:
                            bit_list[((temp1 - temp2) * 8) + j + temp3] = "    " + temp_bitfield_type + " " + \
                                                                          bit.name + " : 1;\n"

            # compress non-dummy bit entries into a single multi-bit entry
            for i in range((address_delta * 8) - 1):
                regpattern = re.compile(r"(.+: )(\d{1,2});")
                regpattern2 = re.compile(r".+\W+(.+):")

                rp1 = regpattern2.search(bit_list[i]).group(1)
                rp2 = regpattern2.search(bit_list[i + 1]).group(1)

                if rp1 == rp2:
                    temp = regpattern.search(bit_list[i])

                    bit_list.pop(i)
                    bit_list.insert(i, "derp")
                    bit_list.pop(i + 1)
                    bit_list.insert(i + 1, temp.group(1) + str(int(temp.group(2)) + 1) + ";\n")

            for row in bit_list:
                if row != "derp":
                    bit_type_string += row
                    regpattern = re.compile(r".* (\S+) : .*")
                    temp = regpattern.search(str(row)).group(1)
                    if not temp.__contains__("padding"):
                        bit_type_macro_string += "#define " + base_type_struct_name[1:] + temp + " " + \
                                                 base_type_struct_name[1:] + "_BASE.BIT." + temp + "\n"
            # Takes off "_" at the front of the word using [1:]
            bit_type_string += "} " + base_type_struct_name[1:] + "Bits_t;\n\n"

            # print bit_type_string
            output_strings.groups[output_group_name].bits += bit_type_string
            output_strings.groups[output_group_name].macros += bit_type_macro_string

        for row in filtered_list:
            excel_rows_sorted.remove(row)

        # if there's anything left in the list, keep parsing
        if excel_rows_sorted:
            address_for_match = excel_rows_sorted[0].address
        else:
            address_for_match = 'DEADBEEF'

    else:
        print "attempting to match address with no entries near 0x" + address_for_match
        exit()

    t1 = time.clock()
    # print "one loop: %f\n" % (t1-t0)

    return address_for_match


def validate_arguments():
    number_of_groups_cnt = 0
    input_file_name_cnt = 0
    tab_name_cnt = 0
    output_file_location_cnt = 0
    base_type_byte_cnt = 0
    base_type_short_cnt = 0
    base_type_long_cnt = 0
    base_union_name_byte_cnt = 0
    base_union_name_short_cnt = 0
    base_union_name_long_cnt = 0
    base_union_name_bits_cnt = 0
    gen_address_macros_cnt = 0
    use_module_name = False
    config_file_valid_lines = []

    if len(sys.argv) < 2:
        print 'syntax error: use CLI "python headerGen.py <config file>"'
        exit()

    config_file_object = open(sys.argv[1], "r")

    # strip blank lines
    config_file_lines = filter(None, (line.rstrip() for line in config_file_object))

    for line in config_file_lines:
        if not line.startswith("#"):
            config_file_valid_lines.append(line)

    for line in config_file_valid_lines:
        if "[groups]" in line:
            number_of_groups_cnt += 1
        if "use_module_name = true" in line or "use_module_name = True" in line:
            use_module_name = True
        if "input_file_name" in line:
            input_file_name_cnt += 1
        if "tab_name" in line:
            tab_name_cnt += 1
        if "output_file_location" in line:
            output_file_location_cnt += 1
        if "base_type_byte" in line:
            base_type_byte_cnt += 1
        if "base_type_short" in line:
            base_type_short_cnt += 1
        if "base_type_long" in line:
            base_type_long_cnt += 1
        if "base_union_name_byte" in line:
            base_union_name_byte_cnt += 1
        if "base_union_name_short" in line:
            base_union_name_short_cnt += 1
        if "base_union_name_long" in line:
            base_union_name_long_cnt += 1
        if "base_union_name_bits" in line:
            base_union_name_bits_cnt += 1
        if "gen_address_macros = " in line:
            gen_address_macros_cnt += 1

    if number_of_groups_cnt is not 0 and use_module_name is True:
        print 'config file syntax error: cannot use "use_module_name" argument with other user-defined [group] \
entries'
        exit()

    if use_module_name is False and number_of_groups_cnt is 0:
        print 'config file syntax error: you must specify at least one [group] argument, or "use_module_name" argument'
        exit()

    if input_file_name_cnt is not 1:
        print 'config file syntax error: ensure a single "input_file_name" argument in config file'
        exit()

    if tab_name_cnt is not 1:
        print 'config file syntax error: ensure a single "tab_name" argument in config file'
        exit()

    if output_file_location_cnt is not 1:
        print 'config file syntax error: ensure a single "output_file_location" argument in config file'
        exit()

    if base_type_byte_cnt is not 1 or base_type_short_cnt is not 1 or base_type_long_cnt is not 1:
        print 'config file syntax error: ensure one base type entry for byte, short, and long types'
        exit()

    if base_union_name_byte_cnt is not 1 or base_union_name_short_cnt is not 1 or base_union_name_long_cnt is not 1 \
            or base_union_name_bits_cnt is not 1:
        print 'config file syntax error: ensure one base union name entry for byte, short, long, and bit members'
        exit()

    if gen_address_macros_cnt is not 1:
        print 'config file syntax error: ensure a single "gen_address_macros" argument in config file'
        exit()

    # done with file object - data lives in config_file_lines now
    config_file_object.close()


# -----------------------------------------------------------------------------------------------
# MAIN()
# -----------------------------------------------------------------------------------------------
def main():
    if len(sys.argv) < 2:
        print 'syntax: python headerGen.py <config file>'
        exit()

    print "running HeaderGen.py version 1.3...\n"

    validate_arguments()

    # pull config file name from command line
    config_file_name = sys.argv[1]

    # pull config data from config file
    config_data = parse_config_file(config_file_name)

    # mask simple warnings for a few lines to avoid warning about Excel file properties - this will clutter the CLI
    warnings.simplefilter("ignore")
    # pull Excel file into list
    excel_rows = []
    max_rows = load_excel(config_data, excel_rows)
    warnings.simplefilter("always")

    # determine if the user wants to use group names from the Excel or not
    if config_data["use_module_name"] is True:
        # read all module names from header file and add them to config_data
        find_module_names(excel_rows, config_data)

    # create one group class for each of the groups specified in the config file
    output_string_groups = {}
    for i in config_data["groups"]:
        output_string_groups[i[0].strip()] = OutputStringGroup("", "", "", "", "")
        output_strings = OutputStrings(config_data["prefix"], output_string_groups, config_data["suffix"])

    # sort array by address
    excel_rows_sorted = sorted(excel_rows, key=sort_by_address)

    # filter list to match a single address
    # result: all entries matching 'address_for_match'
    address_for_match = excel_rows_sorted[0].address
    max_address = excel_rows_sorted[-1].address

    i = 0
    print "\n"
    while address_for_match is not 'DEADBEEF':
        address_for_match = process_register(address_for_match, excel_rows_sorted, output_strings, config_data)
        i+=1
        sys.stdout.flush()
        temp2 = float(i) / float(max_rows) * float(5)
        sys.stdout.write('\rgenerating header file strings: {:.0%} at address 0x{}'.format(temp2, address_for_match))
    print "\n"

    for temp_group_list in config_data["groups"]:
        temp_group = temp_group_list[0]
        output_file = open(temp_group + "_RegDefns.h", "w")

        # print inclusion guard
        if temp_group == 'DEFAULT':
            output_file.write("#ifndef __" + config_data["device_name"].upper() + "_HEADER__\n")
            output_file.write("#define __" + config_data["device_name"].upper() + "_HEADER__\n")
        else:
            output_file.write("#ifndef __" + config_data["device_name"].upper() + "_" + temp_group + "_HEADER__\n")
            output_file.write("#define __" + config_data["device_name"].upper() + "_" + temp_group + "_HEADER__\n")

        output_file.write(output_strings.prefix + "\n")
        output_file.write("/* ------------------------------------------------------------------ */\n")
        output_file.write("/* BIT ACCESS STRUCTURES                                              */\n")
        output_file.write("/* ------------------------------------------------------------------ */\n")
        output_file.write(output_strings.groups[temp_group].bits)
        output_file.write("/* ------------------------------------------------------------------ */\n")
        output_file.write("/* BASE TYPE STRUCTURES                                               */\n")
        output_file.write("/* ------------------------------------------------------------------ */\n")
        output_file.write(output_strings.groups[temp_group].base)
        output_file.write("/* ------------------------------------------------------------------ */\n")
        output_file.write("/* MEMORY MAPPING PRAGMAS                                             */\n")
        output_file.write("/* ------------------------------------------------------------------ */\n")
        output_file.write(output_strings.groups[temp_group].bind + "\n")
        output_file.write("/* ------------------------------------------------------------------ */\n")
        output_file.write("/* REGISTER ACCESS MACROS                                             */\n")
        output_file.write("/* ------------------------------------------------------------------ */\n")
        output_file.write(output_strings.groups[temp_group].macros + "\n")
        if config_data["gen_address_macros"] is True:
            output_file.write("/* ------------------------------------------------------------------ */\n")
            output_file.write("/* REGISTER ADDRESS ACCESS MACROS                                     */\n")
            output_file.write("/* ------------------------------------------------------------------ */\n")
            output_file.write(output_strings.groups[temp_group].macros_addr + "\n")
        output_file.write(output_strings.suffix + "\n")
        output_file.write("#endif")
        output_file.close()

# run main with execution timing
t_start = time.clock()
main()
t_end = time.clock()
print "total run time is: %f seconds" % (t_end-t_start)
